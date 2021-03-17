import streamlit as st
import plotly.express as px
import plotly.graph_objs as go
import plotly.io as pio
import pandas as pd
from openpyxl import load_workbook
import openpyxl

st.write('# 네이버 쇼핑 랭크')
keyword = ['장뇌삼', '장뇌삼 효천영농조합']
file_name = '가비 단일 장뇌삼'
# https://coinmarketcap.com
# scraper = CmcScraper('BTC', '01-01-2021', '07-01-2021')  # '%d-%m-%Y'

#시트이름 불러오기
wb=openpyxl.load_workbook("{}.xlsx".format(file_name))
ws=wb.sheetnames

sheets=[]
box=[]
for i in ws:
    sheets.append(i)
#시트 찾아서 시트의 데이터 가져오기
for i in sheets:
    box.append(pd.read_excel("{}.xlsx".format(file_name), sheet_name=i))
#fig 시작
pio.renderers.default = "notebook_connected"
fig = go.Figure()
try:
    for i in range(0, len(box)):
        for j in box[i]['키워드'].unique():
            fig.add_trace(go.Scatter(x=box[i]['점검시간'],y=box[i]['현재 순위'], name=j))
            fig.update_traces(mode="markers+lines", hovertemplate=None)
            fig.update_layout(hovermode="x")
        fig.update_yaxes(autorange="reversed")
        fig.update_layout(
            title="파일명: {}".format(file_name),
            yaxis_title="현재 순위",
            xaxis_title="점검 시간",
            # width=900,
            # height=800
        )
    #저장
    fig.write_html('{}.html'.format(file_name))
    print('저장완료')
    fig.show()
    st.plotly_chart(fig)
except KeyError as e:
    print('{} 파일이 존재하지 않습니다.'.format(file_name))


# for n in sheet:
#     df = pd.read_excel('{}.xlsx'.format(file_name),sheet_name=n)
#
#     for i in keyword:
#
#         fig_close = px.line(df, x='점검시간', y='현재 순위', title=n)
#     st.plotly_chart(fig_close)


# C:\Users\ysn39\PycharmProjects\pythonProject\streamlit
# streamlit run stream.py
# https://navershoping.df.r.appspot.com