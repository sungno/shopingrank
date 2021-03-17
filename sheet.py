# C:\Users\ysn39\Documents\카카오톡 받은 파일\

# 1. keyword input
import threading
import pandas as pd
import time
from selenium import webdriver
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl import Workbook
from threading import Thread, Lock
import logging

path = input('path 입력:')
interval = int(input('Interval-초단위로 입력: '))
file = pd.read_excel('키워드 관리.xlsx')
pd.options.display.float_format = '{:}'.format  # 소수점 없애고
file = file.convert_dtypes()  # 데이터 타입변경
keyword = file['시트명']
keybox = []
for i in keyword:
    keybox.append(i)

# 파일 없다면
file_name = file.loc[:, '파일명'].unique()
for i in file_name:
    if os.path.exists(r"{}{}.xlsx".format(path, i)):
        pass
    else:
        wb = Workbook()  # 파일 만듬
        wb.save('{}.xlsx'.format(i))


class worker(threading.Thread):
    # threading에서 내장 모듈인 Thread를 상속
    def __init__(self, name):
        super().__init__()
        self.name = name

    def run(self):
        file_name = file.loc[num, '파일명']
        status = file.loc[num, 'Status']
        if status == 0 and '단독' in str(file_name):
            pid = str(file.loc[num, 'P-ID'])
            cid = '-'
            code = pid
        elif status == 0 and '가비' in str(file_name):
            cid = str(file.loc[num, 'C-ID'])
            pid = str(file.loc[num, 'P-ID'])
            code = cid
        else:
            pass
        name = threading.currentThread().getName()

        print('process 시작. {} {}: {}'.format(file_name, name, code))
        page = 80
        columns = ["점검시간", "키워드", "CID", "PID", "현재 순위", "발견페이지", "발견위치", "페이지당 상품 수"]

        wb = openpyxl.load_workbook(r"{}{}.xlsx".format(path, file_name))  # 파일 가져와
        ws = wb.sheetnames

        # 지금 키워드의 시트 없다면
        if 'Sheet' in ws:
            sheet = wb.active
            sheet.title = name  # 시트명 변경
            sheet.append(columns)  # 열데이터 추가
            wb.save(r'{}{}.xlsx'.format(path, file_name))  # 엑셀파일저장
        elif 'Sheet' not in ws and name not in ws:
            sheet = wb.create_sheet()  # 새시트 만들기
            sheet.title = name  # 시트명 변경
            sheet.append(columns)  # 열데이터 추가
            wb.save(r'{}{}.xlsx'.format(path, file_name))  # 엑셀파일저장
        else:
            pass

        try:
            while True:
                quit = 0
                times = time.strftime("%y/%m/%d %H:00")
                driver = webdriver.Chrome('c:/informs/chromedriver.exe')

                prebox = []
                for n in range(1, 10000):

                    sem.acquire()

                    driver.get(
                        "https://search.shopping.naver.com/search/all?frm=NVSHATC&pagingIndex={}&pagingSize={}&productSet=total&query={}&sort=rel&timestamp=&viewType=list".format(
                            n, page, name))
                    driver.implicitly_wait(60)
                    time.sleep(5)
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    driver.implicitly_wait(60)
                    time.sleep(10)

                    source = driver.page_source
                    html = BeautifulSoup(source, "lxml")
                    lists = html.find('ul', attrs={'class': 'list_basis'}).find_all('li', attrs={
                        'class': 'basicList_item__2XT81'})
                    pronums = html.find('a', attrs={'class': 'subFilter_filter__3Y-uy'}).find_all('span', attrs={
                        'class': 'subFilter_num__2x0jq'})
                    for i in pronums:
                        pronums = i.text
                        pronums = pronums.replace(',', '')
                    if int(pronums) >= page:
                        while len(lists) < page:
                            time.sleep(5)
                    else:
                        pass
                    sem.release()

                    afterbox = []
                    for i in lists:
                        if 'N=a:lst*A' in str(i):
                            pass
                        else:
                            afterbox.append(i.find('a')['data-nclick'])
                    sem.release()

                    prebox.append(afterbox)
                    for idx, i in enumerate(afterbox):
                        if str(code) in i:
                            index = idx + 1
                            rank = int(page) * (n - 1) + index
                            datas = [times, name, cid, pid, rank, n, index, page]
                            print("[{}] {}, {} found".format(times, file_name, name))

                            sem.acquire()

                            wb = openpyxl.load_workbook(r"{}{}.xlsx".format(path, file_name))  # 파일 가져와
                            sheet = wb[name]
                            sheet.append(datas)
                            wb.save(r'{}{}.xlsx'.format(path, file_name))  # 엑셀파일저장
                            sem.release()

                            driver.close()
                            quit = 1
                            break
                    if quit == 1:
                        break
                    if len(prebox) > 2:
                        if prebox[-2] == afterbox:
                            print('[{}] {}, {}: not found.'.format(times, file_name, name))
                            driver.close()
                            break
                        else:
                            continue

                wb.save(r'{}{}.xlsx'.format(path, file_name))  # 엑셀파일저장
                time.sleep(interval)

        except AttributeError as e:
            wb.save(r'{}{}.xlsx'.format(path, file_name))
            print('[{}] {}, {}: not found. {}'.format(times, file_name, name, e))
            driver.close()
            pass
            time.sleep(interval)


print('{}스레드 시작'.format(threading.currentThread().getName()))
lock = Lock()
sem = threading.Semaphore(1)
lock.acquire()
num = 0
for i in file.index:
    name = file.loc[i, '시트명']
    t = worker(name)
    num = i
    t.start()
    time.sleep(10)
lock.release()

# C:\Users\c\파이썬 김채연 파일\ranks\
# C:\Users\ysn39\파이썬 주피터\파이썬 진행중\네이버 쇼핑 랭킹\